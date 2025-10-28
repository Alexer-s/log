import os
import pandas as pd
import re
import sys
import tkinter as tk
from datetime import date
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from typing import Union
from tkinter import filedialog, messagebox, ttk
from main_klass import DatabaseError, OutQueryDb, UniqueError


FIELDS_WIN_ADD: dict[str, Union[ttk.Entry, ttk.Combobox]] = {}
# path_to_bd = os.getenv('PATH_TO_BD', None)


names_fields = [
    ('ID', 'id', 25),
    ('Дата рег-ции', 'date_reg', 50),
    ('Номер з-cа:', 'num_query', 40),
    ('Дата запроса:', 'date_query', 80),
    ('Адресат - ТО СФР:', 'to_sfr', 100),
    ('СНИЛС:', 'snils', 88),
    ('Ф.И.О.:', 'fio', 120),
    ('Название орг-ции, рег.№:', 'name_org', 120),
    ('Период работы:    ', 'work_period', 130),
    ('Срок исполнения:', 'due_date', 80),
    ('Вид запроса:', 'type_query', 50),
    ('Дата ответа:', 'date_answer', 80),
    ('Специалист', 'spec', 100)
]


def write_query(
        add_win: tk.Toplevel, update: bool, id: Union[int, None] = None
):
    '''Записывает запрос в БД или обновляет существующий.'''
    if not update:
        FIELDS_WIN_ADD['date_reg'].config(state='normal')
        FIELDS_WIN_ADD['date_reg'].insert(0, date.today().strftime('%d.%m.%Y'))
        FIELDS_WIN_ADD['date_reg'].config(state='disabled')
    params = tuple(val.get() for val in FIELDS_WIN_ADD.values())
    try:
        if update:
            res = gur_query.insert_to_db(
                gur_query.update_query, params + (id,)
            )
        else:
            res = gur_query.insert_to_db(
                gur_query.insert_query, params
            )
    except DatabaseError:
        messagebox.showerror(
            'Ошибка', 'Что-то пошло не так, попробуйте еще раз!'
        )
        return
    if res:
        try:
            fields = gur_query.read_db(
                gur_query.read_query_fo_id, (res if not update else id,)
            )
        except DatabaseError:
            messagebox.showerror('Ошибка', 'При чтении БД возникла ошибка.')
        confirm_win = tk.Toplevel()
        confirm_win.title('Подтверждение')
        confirm_frame = ttk.Frame(
            confirm_win, borderwidth=1, relief=tk.SOLID, padding=(3, 3)
        )
        confirm_frame.pack()
        res_text = ttk.Label(
            confirm_frame,
            text='Вы добавили запрос №:'
            if not update
            else 'Вы обновили запрос №:',
            width=30
        )
        res_text.grid(column=0, row=0, padx=2, pady=3, sticky='w')
        num = ttk.Label(confirm_frame, text=fields[0][0], width=30)
        num.grid(column=1, row=0, padx=2, pady=3, sticky='w')
        for num_row, text, val in zip(
            range(1, len(fields[0])), names_fields[1:], fields[0][1:]
        ):
            label = ttk.Label(confirm_frame, text=text[0], width=30)
            label.grid(column=0, row=num_row, padx=2, pady=3, sticky='w')
            value = ttk.Label(confirm_frame, text=val, width=30)
            value.grid(column=1, row=num_row, padx=2, pady=3, sticky='w')
        add_win.destroy()


# сделать валидацию полей (даты, снилса и специалиста)
def date_validation(event, snils=False):
    '''Ограничивает длину даты, снилса, расставляет разделители.'''
    special_keys = {'Return', 'Control_L', 'Control_R', 'Alt_L', 'Alt_R',
                    'Left', 'Right', 'Up', 'Down', 'Shift_L', 'Shift_R',
                    'Tab', 'Escape', 'Delete', 'BackSpace'}
    if snils:
        max_count, sep = (10, {3: '-', 6: '-', 9: ' '})
    else:
        max_count, sep = (7, {2: '.', 4: '.'})
    if event.keysym not in special_keys:
        widget_val = event.widget.get()
        widget_val_digits = [c for c in widget_val if c.isdigit()]
        if (len(widget_val_digits) > max_count
                and event.keysym not in special_keys):
            return 'break'
        valid_value = ''
        for num, char in enumerate(widget_val_digits):
            if num in sep:
                valid_value += sep[num]
            valid_value += char
            if num >= max_count:
                break
        event.widget.delete(0, tk.END)
        event.widget.insert(0, valid_value)


def win_add_view_query(update=False, update_fields=None):
    '''Формирует окно ввода данных запроса. Если это обновление,
    то вставляет данные запроса.
    '''
    add_win = tk.Toplevel()
    add_win.title('Добавление запроса')
    add_win.geometry('420x410')
    input_frame = ttk.Frame(
        add_win, borderwidth=1, relief=tk.SOLID, padding=(3, 3)
    )
    input_frame.pack(side='left', padx=10, pady=5, anchor='center')
    num_label = ttk.Label(input_frame, text=names_fields[0][0], width=30)
    num_label.grid(column=0, row=0, padx=2, pady=3, sticky='w')
    num_field = ttk.Entry(input_frame, width=30, state='readonly')
    num_field.grid(column=1, row=0, padx=2, pady=3, sticky='w')
    date_reg_label = ttk.Label(input_frame, text=names_fields[1][0], width=30)
    date_reg_label.grid(column=0, row=1, padx=2, pady=3, sticky='w')
    date_reg = ttk.Entry(
        input_frame, width=30, state='readonly'
    )
    FIELDS_WIN_ADD[names_fields[1][1]] = date_reg

    date_reg.grid(column=1, row=1, padx=2, pady=3, sticky='w')
    for line, name_field in zip(
        range(2, len(names_fields)), names_fields[2:-1]
    ):
        label = ttk.Label(input_frame, text=name_field[0], width=30)
        label.grid(column=0, row=line, padx=2, pady=3, sticky='w')
        field = ttk.Entry(input_frame, width=30)
        if name_field[1] in {'date_query', 'due_date', 'date_answer', 'snils'}:
            field.bind(
                '<KeyPress>',
                lambda event, snils=name_field[1] == 'snils': date_validation(
                    event, snils
                )
            )
        FIELDS_WIN_ADD[name_field[1]] = field
        field.grid(column=1, row=line, padx=2, pady=3, sticky='w')
    try:
        users = sorted([spec[1] for spec in gur_query.users_list()])
    except DatabaseError:
        messagebox.showerror(
            'Ошибка', 'Ошибка при получении списка пользователей.'
        )
    spec_field_label = ttk.Label(
        input_frame, text=names_fields[-1][0], width=30
    )
    spec_field_label.grid(
        column=0, row=len(names_fields), padx=2, pady=2, sticky='w'
    )
    print(users)
    spec_field = ttk.Combobox(input_frame, width=30, values=users, state='readonly')
    FIELDS_WIN_ADD[names_fields[-1][1]] = spec_field
    spec_field.grid(
        column=1, row=len(names_fields), padx=2, pady=3, sticky='w'
    )
    if update:
        num_field.config(state='normal')
        num_field.insert(0, update_fields[0])
        num_field.config(state='disabled')
        date_reg.config(state='normal')
        date_reg.insert(0, update_fields[1])
        date_reg.config(state='disabled')
        for field, val in zip(
            list(FIELDS_WIN_ADD.values())[1:], update_fields[2:-1]
        ):
            field.insert(0, val)
        spec_field.set(update_fields[-1])
    add_btn = ttk.Button(
        input_frame,
        text='Добавить'
        if not update
        else 'Обновить',
        command=lambda: write_query(
            add_win, update, id=None if not update else update_fields[0])
        )
    add_btn.grid(
        column=1,
        row=len(names_fields) + 1,
        sticky='e',
        ipadx=3,
        ipady=3,
        padx=3,
        pady=3
    )


def chek_fio(add_user_field):  # разобраться здесь
    pattern = r'^[А-ЯЁ][а-яё]+ [А-ЯЁ].[А-ЯЁ].$'
    # print(f'И сюда глянь {re.match(pattern, "Киселева К.Е.")}')
    if not re.match(pattern, add_user_field.get()):
        add_user_field.config(style='Red.TEntry')
        add_user_field.focus_set()
        return False
    else:
        add_user_field.config(style='TEntry')
        return True


def chek_style(event):
    # print(event.widget.config)
    if event.widget.cget('style') != 'TEntry':
        event.widget.config(style='TEntry')


def create_delete_user(del_usr=False):
    add_del_user_win = tk.Toplevel()
    if del_usr: #  сделать красиво
        add_del_user_win.title('Выбор имени')
        add_del_user_win.geometry('200x300')
        try:
            users = gur_query.read_db(gur_query.all_users)
        except DatabaseError:
            messagebox.showerror(
                'Ошибка', 'Ошибка при чтении БД, попробуйте позже.'
            )
        cols = ['id', 'name']
        tree = ttk.Treeview(
            add_del_user_win,
            columns=cols,
            show='headings'
        )
        tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        tree.heading('id', text='№')
        tree.column('id', width=25, stretch=False)
        tree.heading('name', text='Ф.И.О.')
        tree.column('name', width=100)
        for user in users:
            tree.insert('', tk.END, values=user)
        scrollbar = ttk.Scrollbar(tree, orient='vertical', command=tree.yview)
        scrollbar.pack(side='right', fill='y')
        tree.configure(yscrollcommand=scrollbar.set)
    else:
        add_del_user_win.title('Ввод имени')
        add_del_user_win.geometry('250x150')
        add_user_label = ttk.Label(
            add_del_user_win, text='''Введите имя пользователя по образцу:
            \n"Фамилия И.О."'''
        )
        add_user_label.pack(pady=10)
        add_user_field = ttk.Entry(add_del_user_win, width=30)
        add_user_field.pack(pady=10)
        add_user_field.bind('<KeyPress>', chek_style)

        #  добавить валидацию вводимого значения по образцу "Фамилия И.О."

    def add_user():
        if chek_fio(add_user_field):
            try:
                res = gur_query.insert_to_db(
                    gur_query.create_user,
                    (add_user_field.get(),)
                )
                add_del_user_win.destroy()
                if res:
                    messagebox.showinfo(
                        'Сообщение', 'Пользователь успешно добавлен.'
                    )
            except UniqueError:
                messagebox.showerror(
                    'Ошибка', 'Похоже такой пользователь уже есть.'
                )
            except DatabaseError:
                messagebox.showerror(
                    'Ошибка', 'Что-то пошло не так, попробуйте еще раз.'
                )

    def del_user():
        user_id = tree.item(tree.focus())['values'][0]
        try:
            res = gur_query.insert_to_db(
                gur_query.delete_user,
                (user_id,)
            )
        except DatabaseError:
            messagebox.showerror(
                'Ошибка', 'Что-то пошло не так, попробуйте еще раз.'
            )
        add_del_user_win.destroy()
        if res:
            messagebox.showinfo('Сообщение', 'Пользователь успешно удален.')
        else:
            messagebox.showinfo('Сообщение', 'Нет такого пользователя!')

    add_user_btn = ttk.Button(
        add_del_user_win,
        text='Добавить' if not del_usr else 'Удалить',
        command=add_user if not del_usr else del_user
    )
    add_user_btn.pack(side='right', fill='x', expand=True, padx=5, pady=(0, 5))


def on_close_qur_win():
    gur_query.close_db()
    gur_win.destroy()


def get_query_from_list(event, tree):
    item_id = tree.focus()
    if not item_id:
        return
    values = tree.item(item_id)['values']
    try:
        query = gur_query.read_db(gur_query.read_query_fo_id, (values[0],))
    except DatabaseError:
        messagebox.showerror(
            'Ошибка', 'Ошибка при чтении БД, попробуйте позже.'
        )
    win_add_view_query(update=True, update_fields=query[0])


def find_query(snils):
    try:
        queries = gur_query.read_db(gur_query.read_query, (snils,))
    except DatabaseError:
        messagebox.showerror(
            'Ошибка', 'Ошибка при чтении БД, попробуйте позже.'
        )
    if queries:
        len_queries = len(queries)
        if len_queries == 1:
            win_add_view_query(update=True, update_fields=queries[0])
        if len_queries > 1:
            win_for_many_queries = tk.Toplevel()
            cols = [col[1] for col in names_fields]
            tree = ttk.Treeview(
                win_for_many_queries,
                columns=cols,
                show='headings'
            )
            tree.pack(fill=tk.BOTH, expand=True)
            for name, col, width_col in names_fields:
                tree.heading(col, text=name)
                tree.column(col, width=width_col)
            for query in queries:
                tree.insert('', tk.END, values=query)
            tree.bind(
                '<Double-1>',
                lambda event, tree=tree: get_query_from_list(event, tree)
            )
    else:
        messagebox.showinfo('Сообщение', 'Запрос не найден!')


def out_to_excel(spec):
    try:
        data = pd.read_sql_query(
            gur_query.read_db_query if spec == 'Все' else gur_query.read_spec_queries,
            gur_query.con,
            params=() if spec == 'Все' else (spec,)
        )
    except Exception as error:
        messagebox.showerror(
            'Ошибка',
            f'Ошибка чтения БД при формировании Excel файла: {error}'
        )
    rus_columns = [name[0] for name in names_fields]
    data.columns = rus_columns
    if not os.path.exists(r'C:\output_opp'):
        os.makedirs(r'C:\output_opp', exist_ok=True)
    xcl_addr = os.path.join(r'C:\output_opp', 'log.xlsx')
    print(xcl_addr)
    data.to_excel(xcl_addr, index=False)
    wb = load_workbook(xcl_addr)
    ws = wb.active
    widths = [int(field[2] / 100 * 20) for field in names_fields]
    for col_num, column_cells, width in zip(range(1, len(widths) + 1), ws.columns, widths):
        print(get_column_letter(col_num))
        for cell in column_cells:
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
        ws.column_dimensions[get_column_letter(col_num)].width = width
    wb.save(xcl_addr)
    os.startfile(xcl_addr)


def choise_win():  # навести красоту
    choice_cpec_win = tk.Toplevel()
    choice_cpec_win.geometry('200x120')
    choice_cpec_win.title('Формирование журнала')
    choice_label = ttk.Label(choice_cpec_win, text='Выбрать:')
    choice_label.pack(side='top', anchor='nw', padx=10, pady=10)
    specialists = ['Все',] + sorted([spec[1] for spec in gur_query.users_list()])
    choice_box = ttk.Combobox(choice_cpec_win, width=20, values=specialists, state='readonly')
    choice_box.pack(side='top', fill='x', expand=True, padx=10)
    choice_box.set(specialists[0])

    def on_click():
        out_to_excel(choice_box.get())
        choice_cpec_win.destroy()

    choice_btn = ttk.Button(
        choice_cpec_win,
        text='Сформировать',
        command=on_click
    )
    choice_btn.pack(side='top', anchor='se', padx=10, pady=10)


data_path = getattr(sys, '_MEIPASS', 'C:/Dev/myprojects/logs_for_data')
icon_addr = os.path.join(data_path, 'log.ico')
gur_win = tk.Tk()
gur_win.title('Журнал исходящих запросов')
gur_win.geometry('400x300')
gur_win.iconbitmap(default=icon_addr)
gur_win.protocol('WM_DELETE_WINDOW', on_close_qur_win)
db_name = 'db_query.sqlite'
local_appdata = os.getenv('LOCALAPPDATA')
if local_appdata is None:
    local_appdata = 'C:/'
os.makedirs(os.path.join(local_appdata, 'query_log'), exist_ok=True)
env_addr_path = os.path.join(local_appdata, 'query_log', '.env')
load_dotenv(dotenv_path=env_addr_path)
db_addr = os.getenv('PATH_TO_DB')
if not db_addr:
    gur_win.withdraw()
    folder_path = filedialog.askdirectory(title='Выберите папку для БД')
    if not folder_path:
        gur_win.destroy()
        sys.exit()
    with open(env_addr_path, 'a', encoding='utf-8') as file:
        addr = os.path.join(folder_path, db_name).replace('\\', '/')
        file.write(f'PATH_TO_DB={addr}')
    gur_win.deiconify()
    load_dotenv(dotenv_path=env_addr_path)
    db_addr = os.getenv('PATH_TO_DB')
gur_query = OutQueryDb(db_addr)
red_style = ttk.Style()
red_style.configure('Red.TEntry', foreground='red')

main_menu = tk.Menu()
user_menu = tk.Menu(tearoff=0)
user_menu.add_command(label='Добавить', command=create_delete_user)
user_menu.add_command(
    label='Удалить',
    command=lambda: create_delete_user(del_usr=True)
)
main_menu.add_cascade(label='Пользователи', menu=user_menu)

btn_add = ttk.Button(
    gur_win, text='Зарегистрировать', command=win_add_view_query
)
btn_add.pack(anchor='nw', padx=15, pady=(25, 10))
find_frame = ttk.Frame(gur_win)
find_frame.pack(anchor='nw', padx=15)
find_snils = ttk.Entry(find_frame, width=17)
find_snils.pack(side='left', pady=10)
find_snils.bind('<KeyPress>', lambda event: date_validation(event, snils=True))
find_btn = ttk.Button(
    find_frame,
    text='Найти',
    command=lambda: find_query(find_snils.get())
)
find_btn.pack(side='left', padx=10, pady=10)
out_to_excel_btn = ttk.Button(
    gur_win,
    text='Вывод в Excel',
    command=choise_win,
    width=16
)
out_to_excel_btn.pack(side='bottom', anchor='se', padx=15, pady=10)

gur_win.config(menu=main_menu)
gur_win.mainloop()

#  открывать excel файл после формирования
#  навести красоту
